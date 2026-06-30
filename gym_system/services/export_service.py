import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO


def _fmt_date(value, fmt='%d/%m/%Y'):
    if value is None:
        return 'N/A'
    try:
        if hasattr(value, 'strftime'):
            return value.strftime(fmt)
        return str(value)
    except Exception:
        return str(value)


class ExportService:

    @staticmethod
    def _header_style(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="1a1a2e")
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _auto_width(ws, min_width=12):
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                min_width,
                max((len(str(cell.value or '')) for cell in col), default=min_width)
            )

    @staticmethod
    def export_clients_excel(clients):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Clientes'
        ExportService._header_style(ws, [
            'ID', 'Nombre Completo', 'Tipo Doc', 'Número Doc',
            'Email', 'Celular', 'Género', 'Fecha Inscripción', 'Estado',
        ])
        for c in clients:
            ws.append([
                c.id,
                c.full_name,
                c.document_type,
                c.document_number,
                c.email or '',
                c.phone or '',
                c.gender or '',
                str(c.enrollment_date) if c.enrollment_date else '',
                'Activo' if c.is_active else 'Inactivo',
            ])
        ExportService._auto_width(ws)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def export_payments_excel(payments):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Pagos'
        ExportService._header_style(ws, [
            'ID', 'Cliente', 'Membresía', 'Monto',
            'Fecha Pago', 'Inicio', 'Vencimiento', 'Método',
        ])
        for p in payments:
            ws.append([
                p.id,
                p.client.full_name if p.client else 'N/A',
                p.membership.name  if p.membership else 'N/A',
                p.amount,
                _fmt_date(p.payment_date),
                _fmt_date(p.start_date),
                _fmt_date(p.end_date),
                p.payment_method or '',
            ])
        ExportService._auto_width(ws)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def export_sales_excel(sales):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ventas'
        ExportService._header_style(ws, [
            'ID', 'Cliente', 'Total', 'Método Pago', 'Fecha', 'Nº Productos',
        ])
        for s in sales:
            ws.append([
                s.id,
                s.client.full_name if s.client else 'Cliente general',
                s.total,
                s.payment_method or '',
                _fmt_date(s.sale_date, '%d/%m/%Y %H:%M'),
                len(s.items) if s.items else 0,
            ])
        ExportService._auto_width(ws)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def export_expired_excel(payments, today, label='Vencidas'):
        """
        Exporta lista de membresías vencidas o por vencer.
        Incluye nombre, teléfono y email para que el recepcionista
        pueda contactar a los clientes directamente.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = label[:31]  # Excel limita a 31 chars el nombre de hoja

        ExportService._header_style(ws, [
            'Cliente', 'Teléfono', 'Email',
            'Plan', 'Venció / Vence', 'Días',
        ])

        for p in payments:
            dias = (p.end_date - today).days  # negativo = ya venció
            ws.append([
                p.client.full_name if p.client else 'N/A',
                p.client.phone     if p.client else '',
                p.client.email     if p.client else '',
                p.membership.name  if p.membership else 'N/A',
                _fmt_date(p.end_date),
                dias,
            ])

        # Colorear filas por urgencia
        red    = PatternFill("solid", fgColor="FFCCCC")
        yellow = PatternFill("solid", fgColor="FFF3CD")
        for row in ws.iter_rows(min_row=2):
            try:
                dias_val = int(row[5].value)
            except (TypeError, ValueError):
                continue
            fill = red if dias_val < 0 else yellow
            for cell in row:
                cell.fill = fill

        ExportService._auto_width(ws)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
