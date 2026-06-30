"""
import_clientes.py
=================
Importa los 221 clientes del Excel REGISTRO_DIARIO_PRO-GYM_JUNIO.xlsx
a la base de datos del sistema gym_system.

USO (desde la carpeta gym_system/):
    python import_clientes.py --excel /ruta/al/REGISTRO_DIARIO_PRO-GYM_JUNIO.xlsx

REQUISITOS:
    - Tener las variables de entorno configuradas (.env o Render)
    - openpyxl instalado (ya está en requirements.txt)

LO QUE HACE:
    1. Lee la hoja REGISTRO del Excel
    2. Crea los planes de membresía faltantes (mensual, quincenal, semana)
    3. Importa los 221 clientes (el duplicado Carlos Adam Duarte se crea una sola vez)
    4. Crea el pago/membresía histórico con fecha de vencimiento real
    5. Marca como is_migrated=True para identificar clientes importados
    6. Imprime un resumen al final
"""

import sys
import os
import argparse
from datetime import datetime, date, timedelta
import pytz

# ── Asegurarse de que el path incluye el proyecto ──────────────────────
# Ejecutar siempre desde la carpeta gym_system/
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from dotenv import load_dotenv
load_dotenv()

import openpyxl
from app import application
from database.db import db
from database.models.client import Client
from database.models.membership import Membership
from database.models.payment import Payment

BOGOTA = pytz.timezone('America/Bogota')

# ── Mapeo de planes del Excel → membership_type del sistema ───────────
PLAN_MAP = {
    'mensual':    {'type': 'mensual',    'name': 'Mensual',    'days': 30,  'price': 60000},
    'quincenal':  {'type': 'mensual',    'name': 'Quincenal',  'days': 15,  'price': 35000},
    'semana':     {'type': 'mensual',    'name': 'Semanal',    'days': 7,   'price': 20000},
    # fallback por si aparece otro valor
    'default':    {'type': 'mensual',    'name': 'Mensual',    'days': 30,  'price': 60000},
}


def get_or_create_membership(plan_key: str) -> Membership:
    """Devuelve la membresía existente o la crea si no existe."""
    cfg = PLAN_MAP.get(plan_key.lower(), PLAN_MAP['default'])
    m = Membership.query.filter_by(name=cfg['name']).first()
    if not m:
        m = Membership(
            name=cfg['name'],
            membership_type=cfg['type'],
            duration_days=cfg['days'],
            price=cfg['price'],
            max_members=1,
            requires_student=False,
            is_active=True,
        )
        db.session.add(m)
        db.session.flush()   # obtener id sin commit
        print(f"  [+] Membresía creada: {cfg['name']} ({cfg['days']} días)")
    return m


def parse_date(val) -> date | None:
    """Convierte datetime de openpyxl o string a date."""
    if val is None:
        return None
    if hasattr(val, 'date'):
        return val.date()
    if isinstance(val, date):
        return val
    # intento string
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            pass
    return None


def normalize_phone(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip().replace('.0', '')
    return s if s else None


def run(excel_path: str):
    print(f"\n📂 Leyendo: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if 'REGISTRO' not in wb.sheetnames:
        print("❌ No se encontró la hoja REGISTRO en el Excel.")
        sys.exit(1)

    ws = wb['REGISTRO']
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # ── Filtrar filas vacías ────────────────────────────────────────────
    data = []
    for r in rows:
        nombre = r[0]
        if nombre and str(nombre).strip():
            data.append(r)

    print(f"📋 Filas con datos: {len(data)}")

    with application.app_context():
        creados   = 0
        omitidos  = 0   # ya existían
        errores   = 0
        vistos    = {}  # nombre_normalizado → True  (para deduplicar)

        for r in data:
            nombre_raw   = str(r[0]).strip()
            telefono_raw = r[1]
            fecha_insc   = parse_date(r[2])
            plan_raw     = str(r[3]).strip().lower() if r[3] else 'mensual'
            vencimiento  = parse_date(r[4])
            # estado     = r[5]  (no lo usamos, lo calculamos por fecha)
            # nota       = r[6]

            # ── Deduplicar por nombre normalizado ──────────────────────
            nombre_key = nombre_raw.lower().strip()
            if nombre_key in vistos:
                print(f"  [DUP] Ignorado duplicado: {nombre_raw}")
                omitidos += 1
                continue
            vistos[nombre_key] = True

            # ── Verificar si ya existe en la BD por nombre ─────────────
            existe = Client.query.filter(
                db.func.lower(Client.full_name) == nombre_key
            ).first()
            if existe:
                print(f"  [YA EXISTE] {nombre_raw}")
                omitidos += 1
                continue

            try:
                # ── Crear cliente ──────────────────────────────────────
                telefono = normalize_phone(telefono_raw)
                enrollmt = fecha_insc or date.today()
                is_active = (vencimiento >= date.today()) if vencimiento else False

                cliente = Client(
                    full_name       = nombre_raw,
                    document_type   = 'CC',
                    document_number = f'MIGRADO-{nombre_key.replace(" ", "-")[:40]}',
                    phone           = telefono,
                    enrollment_date = enrollmt,
                    is_active       = is_active,
                    is_migrated     = True,
                    notes           = 'Importado desde Excel REGISTRO_DIARIO_PRO-GYM_JUNIO',
                )
                db.session.add(cliente)
                db.session.flush()   # obtener cliente.id

                # ── Crear membresía y pago histórico ───────────────────
                if vencimiento and fecha_insc:
                    membresia = get_or_create_membership(plan_raw)
                    pago = Payment(
                        client_id      = cliente.id,
                        membership_id  = membresia.id,
                        amount         = 0,          # monto histórico desconocido
                        payment_date   = fecha_insc,
                        start_date     = fecha_insc,
                        end_date       = vencimiento,
                        payment_method = 'efectivo',
                        notes          = f'Pago histórico migrado desde Excel. Plan: {plan_raw}',
                        is_deleted     = False,
                    )
                    db.session.add(pago)

                creados += 1
                estado_txt = "✅ activo" if is_active else "⚪ vencido"
                print(f"  [OK] {nombre_raw:<35} | {plan_raw:<12} | vence: {vencimiento} | {estado_txt}")

            except Exception as e:
                db.session.rollback()
                print(f"  [ERROR] {nombre_raw}: {e}")
                errores += 1
                continue

        # ── Commit final ───────────────────────────────────────────────
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"\n❌ Error al guardar en la base de datos: {e}")
            sys.exit(1)

        print(f"""
╔══════════════════════════════════════╗
║         IMPORTACIÓN COMPLETA         ║
╠══════════════════════════════════════╣
║  ✅ Clientes creados   : {creados:<12}║
║  ⚠️  Ya existían/dup.  : {omitidos:<12}║
║  ❌ Errores            : {errores:<12}║
╚══════════════════════════════════════╝
""")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Importar clientes desde Excel')
    parser.add_argument(
        '--excel',
        required=True,
        help='Ruta al archivo REGISTRO_DIARIO_PRO-GYM_JUNIO.xlsx'
    )
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"❌ Archivo no encontrado: {args.excel}")
        sys.exit(1)

    run(args.excel)
